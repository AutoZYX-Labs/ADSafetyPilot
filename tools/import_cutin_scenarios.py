#!/usr/bin/env python3
"""Import DRIVEResearch cut-in Scenario.xlsx files into an event-level CSV."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - depends on runtime packaging.
    raise SystemExit(
        "openpyxl is required to read .xlsx scenario files. "
        "Use the Codex bundled Python runtime or install openpyxl."
    ) from exc


DEFAULT_SPEED_BINS = [0.0, 40.0, 60.0, 80.0, 100.0, 120.0]
OUTPUT_COLUMNS = [
    "event_id",
    "source_id",
    "source_package",
    "source_file",
    "road_type",
    "road_section_type",
    "city",
    "location_label",
    "traffic_condition",
    "collection_time_period",
    "speed_bin",
    "overlap_type",
    "ego_id",
    "cutin_id",
    "reference_frame",
    "reference_time_definition",
    "keyframe1",
    "keyframe2",
    "keyframe3",
    "keyframe4",
    "keyframe5",
    "ego_speed_kmh",
    "cutin_speed_kmh",
    "relative_speed_kmh",
    "longitudinal_gap_m",
    "longitudinal_gap_signed_m",
    "lateral_speed_mps",
    "cutin_angle_deg",
    "cutin_duration_s",
    "ttc_s",
    "thw_s",
    "min_ttc_s",
    "min_thw_s",
    "ego_max_decel_mps2",
    "cutin_lateral_accel_mps2",
    "sample_quality",
    "quality_notes",
]


@dataclass
class VehicleSample:
    vehicle_id: str
    vehicle_class: str
    frame: int
    x: float
    y: float
    heading_deg: float | None
    longitudinal_velocity_mps: float | None
    lateral_velocity_mps: float | None
    longitudinal_acceleration_mps2: float | None
    lateral_acceleration_mps2: float | None
    x_velocity_mps: float | None
    y_velocity_mps: float | None
    dhw_m: float | None
    thw_s: float | None
    ttc_s: float | None


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if math.isnan(number) or math.isinf(number):
            return None
        return number
    text = str(value).strip()
    if not text or text.lower() in {"na", "n/a", "none", "null", "-"}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def parse_int(value: Any) -> int | None:
    number = parse_float(value)
    if number is None:
        return None
    return int(round(number))


def fmt(value: Any, digits: int = 3) -> str:
    number = parse_float(value)
    if number is None:
        return ""
    rounded = round(number, digits)
    if rounded == 0:
        rounded = 0.0
    text = f"{rounded:.{digits}f}".rstrip("0").rstrip(".")
    return text or "0"


def signed_heading_delta(a: float | None, b: float | None) -> float | None:
    if a is None or b is None:
        return None
    return (a - b + 180.0) % 360.0 - 180.0


def speed_bin(speed_kmh: float | None, bins: list[float]) -> str:
    if speed_kmh is None:
        return "unknown"
    for lower, upper in zip(bins, bins[1:]):
        if lower <= speed_kmh < upper:
            return f"{int(lower)}-{int(upper)}"
    if speed_kmh >= bins[-1]:
        return f">={int(bins[-1])}"
    return f"<{int(bins[0])}"


def parse_speed_bins(text: str) -> list[float]:
    bins = [float(part.strip()) for part in text.split(",") if part.strip()]
    if len(bins) < 2:
        raise ValueError("At least two speed-bin boundaries are required.")
    if bins != sorted(bins):
        raise ValueError("Speed-bin boundaries must be sorted ascending.")
    return bins


def sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    columns = [str(value).strip() if value is not None else "" for value in header]
    output = []
    for row in rows:
        record = {}
        for column, value in zip(columns, row):
            if column:
                record[column] = value
        if record:
            output.append(record)
    return output


def workbook_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return sheet_rows(workbook, sheet_name)
    finally:
        workbook.close()


def load_samples(rows: list[dict[str, Any]]) -> dict[int, VehicleSample]:
    samples = {}
    for row in rows:
        frame = parse_int(row.get("frame"))
        x = parse_float(row.get("xPosition"))
        y = parse_float(row.get("yPosition"))
        if frame is None or x is None or y is None:
            continue
        samples[frame] = VehicleSample(
            vehicle_id=str(row.get("id") or ""),
            vehicle_class=str(row.get("class") or ""),
            frame=frame,
            x=x,
            y=y,
            heading_deg=parse_float(row.get("heading")),
            longitudinal_velocity_mps=parse_float(row.get("longitudinal_velocity_m_per_s")),
            lateral_velocity_mps=parse_float(row.get("lateral_velocity_m_per_s")),
            longitudinal_acceleration_mps2=parse_float(row.get("longitudinal_acceleration_m_per_s2")),
            lateral_acceleration_mps2=parse_float(row.get("lateral_acceleration_m_per_s2")),
            x_velocity_mps=parse_float(row.get("xVelocity")),
            y_velocity_mps=parse_float(row.get("yVelocity")),
            dhw_m=parse_float(row.get("DHW")),
            thw_s=parse_float(row.get("THW")),
            ttc_s=parse_float(row.get("TTC")),
        )
    return samples


def feature_row(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return rows[0] if rows else {}


def load_workbook_data(path: Path) -> tuple[dict[int, VehicleSample], dict[int, VehicleSample], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        traffic_rows = sheet_rows(workbook, "traffic_veh")
        ego_rows = sheet_rows(workbook, "ego_veh")
        feature_rows = sheet_rows(workbook, "feature")
    finally:
        workbook.close()
    return load_samples(ego_rows), load_samples(traffic_rows), feature_row(feature_rows)


def choose_reference_frame(
    features: dict[str, Any],
    common_frames: set[int],
    preferred_keys: list[str],
) -> tuple[int | None, str]:
    for key in preferred_keys:
        frame = parse_int(features.get(key))
        if frame in common_frames:
            return frame, key
    candidates = [parse_int(features.get(key)) for key in preferred_keys]
    candidates = [frame for frame in candidates if frame is not None]
    if common_frames and candidates:
        preferred = candidates[0]
        nearest = min(common_frames, key=lambda frame: abs(frame - preferred))
        return nearest, f"nearest_common_frame_to_{preferred_keys[0]}"
    if common_frames:
        return min(common_frames), "first_common_frame"
    return None, "no_common_frame"


def velocity_unit(sample: VehicleSample) -> tuple[float, float] | None:
    vx = sample.x_velocity_mps
    vy = sample.y_velocity_mps
    if vx is None or vy is None:
        return None
    norm = math.hypot(vx, vy)
    if norm <= 1e-9:
        return None
    return vx / norm, vy / norm


def longitudinal_gap(ego: VehicleSample, target: VehicleSample) -> float | None:
    unit = velocity_unit(ego)
    if unit is None:
        return None
    ux, uy = unit
    return (target.x - ego.x) * ux + (target.y - ego.y) * uy


def thw(ego: VehicleSample, target: VehicleSample) -> float | None:
    gap = longitudinal_gap(ego, target)
    speed = ego.longitudinal_velocity_mps
    if gap is None or speed is None or abs(speed) <= 1e-9:
        return None
    return abs(gap) / abs(speed)


def minimum_thw(ego_samples: dict[int, VehicleSample], target_samples: dict[int, VehicleSample]) -> float | None:
    values = []
    for frame in sorted(set(ego_samples) & set(target_samples)):
        value = ego_samples[frame].thw_s
        if value is not None and value >= 0:
            values.append(value)
    return min(values) if values else None


def max_abs_lateral_accel(samples: dict[int, VehicleSample]) -> float | None:
    values = [
        abs(sample.lateral_acceleration_mps2)
        for sample in samples.values()
        if sample.lateral_acceleration_mps2 is not None
    ]
    return max(values) if values else None


def ids_from_filename(path: Path) -> tuple[str, str]:
    match = re.search(r"Ego_(?P<ego>[^_]+)_CutIn_(?P<cutin>[^_]+)_", path.stem)
    if not match:
        return "", ""
    return match.group("ego"), match.group("cutin")


def import_file(path: Path, args: argparse.Namespace, bins: list[float]) -> dict[str, str]:
    ego_samples, cutin_samples, features = load_workbook_data(path)
    common_frames = set(ego_samples) & set(cutin_samples)
    reference_keys = [key.strip() for key in args.reference_keyframes.split(",") if key.strip()]
    reference_frame, reference_key = choose_reference_frame(features, common_frames, reference_keys)
    ego_id, cutin_id = ids_from_filename(path)
    source_id = path.parents[2].name
    event_id = f"{source_id}::{path.stem}"
    notes = []

    if reference_frame is None:
        notes.append("no_common_reference_frame")
        ego_ref = None
        cutin_ref = None
    else:
        ego_ref = ego_samples[reference_frame]
        cutin_ref = cutin_samples[reference_frame]

    ego_speed_mps = ego_ref.longitudinal_velocity_mps if ego_ref else None
    cutin_speed_mps = cutin_ref.longitudinal_velocity_mps if cutin_ref else None
    ego_speed_kmh = abs(ego_speed_mps) * 3.6 if ego_speed_mps is not None else None
    cutin_speed_kmh = abs(cutin_speed_mps) * 3.6 if cutin_speed_mps is not None else None
    relative_speed_kmh = (
        cutin_speed_kmh - ego_speed_kmh
        if cutin_speed_kmh is not None and ego_speed_kmh is not None
        else None
    )
    gap_signed = None
    gap = None
    if ego_ref:
        if ego_ref.dhw_m is not None and ego_ref.dhw_m >= 0:
            gap = ego_ref.dhw_m
        elif ego_ref.dhw_m is not None:
            notes.append("negative_source_dhw")
        if gap is None and cutin_ref:
            gap_signed = longitudinal_gap(ego_ref, cutin_ref)
            gap = abs(gap_signed) if gap_signed is not None else None
            if gap is not None:
                notes.append("computed_gap_from_geometry")
    thw_ref = None
    if ego_ref:
        if ego_ref.thw_s is not None and ego_ref.thw_s >= 0:
            thw_ref = ego_ref.thw_s
        elif ego_ref.thw_s is not None:
            notes.append("negative_source_thw")
        elif gap is not None and ego_speed_mps is not None and abs(ego_speed_mps) > 1e-9:
            thw_ref = abs(gap) / abs(ego_speed_mps)
            notes.append("computed_thw_from_gap")
    min_ttc = parse_float(features.get("TTC_min_s"))
    ttc_ref = ego_ref.ttc_s if ego_ref and ego_ref.ttc_s is not None and ego_ref.ttc_s >= 0 else None
    if ego_ref and ego_ref.ttc_s is not None and ego_ref.ttc_s < 0:
        notes.append("negative_source_ttc")
    min_thw = minimum_thw(ego_samples, cutin_samples)
    keyframe1 = parse_int(features.get(args.duration_start_keyframe))
    keyframe4 = parse_int(features.get(args.duration_end_keyframe))
    keyframe5 = parse_int(features.get("keyframe5"))
    duration = (keyframe4 - keyframe1) / 30.0 if keyframe1 is not None and keyframe4 is not None else None
    ego_max_decel = parse_float(features.get("Max_deacceration_mps2"))
    cutin_lat_accel = max_abs_lateral_accel(cutin_samples)
    angle = signed_heading_delta(
        cutin_ref.heading_deg if cutin_ref else None,
        ego_ref.heading_deg if ego_ref else None,
    )
    lateral_speed = (
        abs(cutin_ref.lateral_velocity_mps)
        if cutin_ref and cutin_ref.lateral_velocity_mps is not None
        else None
    )

    required_values = [ego_speed_kmh, cutin_speed_kmh, gap, lateral_speed, duration, ttc_ref, thw_ref]
    if not ego_samples:
        notes.append("missing_ego_sheet_data")
    if not cutin_samples:
        notes.append("missing_traffic_sheet_data")
    if any(value is None for value in required_values):
        notes.append("missing_required_metric")
    quality = "valid" if not notes else "review"

    return {
        "event_id": event_id,
        "source_id": source_id,
        "source_package": args.source_package,
        "source_file": str(path.relative_to(args.input_dir)),
        "road_type": args.road_type,
        "road_section_type": args.road_section_type,
        "city": args.city,
        "location_label": args.location_label,
        "traffic_condition": args.traffic_condition,
        "collection_time_period": args.collection_time_period,
        "speed_bin": speed_bin(ego_speed_kmh, bins),
        "overlap_type": args.overlap_type,
        "ego_id": ego_id,
        "cutin_id": cutin_id,
        "reference_frame": fmt(reference_frame, 0),
        "reference_time_definition": (
            f"{reference_key}; kinematic metrics, ttc_s, and thw_s use this frame; "
            f"cutin_duration_s uses {args.duration_end_keyframe}-{args.duration_start_keyframe}; "
            "min_ttc_s uses feature.TTC_min_s"
        ),
        "keyframe1": fmt(features.get("keyframe1"), 0),
        "keyframe2": fmt(features.get("keyframe2"), 0),
        "keyframe3": fmt(features.get("keyframe3"), 0),
        "keyframe4": fmt(features.get("keyframe4"), 0),
        "keyframe5": fmt(features.get("keyframe5"), 0),
        "ego_speed_kmh": fmt(ego_speed_kmh),
        "cutin_speed_kmh": fmt(cutin_speed_kmh),
        "relative_speed_kmh": fmt(relative_speed_kmh),
        "longitudinal_gap_m": fmt(gap),
        "longitudinal_gap_signed_m": fmt(gap_signed),
        "lateral_speed_mps": fmt(lateral_speed),
        "cutin_angle_deg": fmt(angle),
        "cutin_duration_s": fmt(duration),
        "ttc_s": fmt(ttc_ref),
        "thw_s": fmt(thw_ref),
        "min_ttc_s": fmt(min_ttc),
        "min_thw_s": fmt(min_thw),
        "ego_max_decel_mps2": fmt(ego_max_decel),
        "cutin_lateral_accel_mps2": fmt(cutin_lat_accel),
        "sample_quality": quality,
        "quality_notes": ";".join(notes),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_dir", type=Path)
    parser.add_argument("output_csv", type=Path)
    parser.add_argument("--summary-json", type=Path)
    parser.add_argument("--source-package", default="DRIVEResearch_expressway_1h_cutin_scenarios_20260416")
    parser.add_argument("--road-type", default="urban_expressway")
    parser.add_argument("--road-section-type", default="straight_or_ramp_area")
    parser.add_argument("--city", default="Changchun")
    parser.add_argument("--location-label", default="Nan Fourth Ring Expressway near Changchun Water Group")
    parser.add_argument("--traffic-condition", default="congested")
    parser.add_argument("--collection-time-period", default="2025-08-02 07:21-08:41")
    parser.add_argument("--overlap-type", default="normal")
    parser.add_argument("--speed-bins", default=",".join(str(int(value)) for value in DEFAULT_SPEED_BINS))
    parser.add_argument(
        "--reference-keyframes",
        default="keyframe3,keyframe2,keyframe1,keyframe4,keyframe5",
        help="Comma-separated feature keyframes to try as the reference frame, in priority order.",
    )
    parser.add_argument("--duration-start-keyframe", default="keyframe1")
    parser.add_argument("--duration-end-keyframe", default="keyframe4")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    bins = parse_speed_bins(args.speed_bins)
    files = sorted(args.input_dir.glob("**/cutin_scenarios/*.xlsx"))
    if not files:
        raise SystemExit(f"No cut-in scenario .xlsx files found under {args.input_dir}")

    rows = [import_file(path, args, bins) for path in files]
    args.output_csv.parent.mkdir(parents=True, exist_ok=True)
    with args.output_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)

    valid_count = sum(1 for row in rows if row["sample_quality"] == "valid")
    speed_bin_counts: dict[str, int] = {}
    for row in rows:
        speed_bin_counts[row["speed_bin"]] = speed_bin_counts.get(row["speed_bin"], 0) + 1

    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "input_dir": str(args.input_dir),
        "output_csv": str(args.output_csv),
        "source_package": args.source_package,
        "file_count": len(files),
        "row_count": len(rows),
        "valid_count": valid_count,
        "review_count": len(rows) - valid_count,
        "speed_bin_counts": dict(sorted(speed_bin_counts.items())),
        "field_notes": {
            "reference_frame": (
                f"{args.reference_keyframes} priority order; otherwise nearest common frame. "
                "The mapping between workbook keyframes and legacy T1/T2 labels must be confirmed before external release."
            ),
            "ttc_s": "ego_veh.TTC at the reference frame; negative or missing source values stay blank and mark the row for review.",
            "min_ttc_s": "feature.TTC_min_s.",
            "thw_s": "ego_veh.THW at the reference frame; negative values stay blank and mark the row for review.",
            "relative_speed_kmh": "cut-in vehicle speed minus ego speed; negative means cut-in vehicle is slower.",
            "longitudinal_gap_m": "ego_veh.DHW at the reference frame when non-negative; otherwise geometric fallback with a quality note.",
            "cutin_angle_deg": "signed heading difference, cut-in vehicle heading minus ego heading, normalized to [-180, 180].",
            "cutin_duration_s": f"({args.duration_end_keyframe} - {args.duration_start_keyframe}) / 30 fps.",
        },
    }
    if args.summary_json:
        args.summary_json.parent.mkdir(parents=True, exist_ok=True)
        args.summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        f"Wrote {args.output_csv} with {len(rows)} rows "
        f"({valid_count} valid, {len(rows) - valid_count} review)."
    )
    if args.summary_json:
        print(f"Wrote {args.summary_json}.")


if __name__ == "__main__":
    main()
